"""
Celery tasks for FCD report generation.
Produces PDF (WeasyPrint) and Excel (openpyxl) reports per AOI/run.
"""
import os
from pathlib import Path
from celery import shared_task
from sqlalchemy import create_engine, text
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
from weasyprint import HTML

DATABASE_URL = os.environ["DATABASE_URL"]
REPORT_DIR   = Path(os.getenv("REPORT_OUTPUT_DIR", "/data/reports"))
BASE_URL      = os.getenv("PUBLIC_BASE_URL", "http://localhost:8000")

engine = create_engine(DATABASE_URL.replace("+asyncpg", ""))

CLASS_COLORS = {
    0: "000000",
    1: "3380cc",
    2: "006837",
    3: "3ea540",
    4: "baf096",
    5: "ad8855",
}


def _hex(code: int) -> str:
    return "#" + CLASS_COLORS.get(code, "cccccc")


def _fetch_stats(run_id: str, aoi_id: str) -> dict:
    with engine.connect() as conn:
        aoi = conn.execute(
            text("SELECT name, unit_type, district FROM admin_units WHERE id = :id"),
            {"id": aoi_id},
        ).fetchone()

        classes = conn.execute(
            text(
                """SELECT class_code, class_name, area_ha, percent_area
                   FROM fcd_class_stats
                   WHERE run_id = :r AND aoi_id = :a
                   ORDER BY class_code"""
            ),
            {"r": run_id, "a": aoi_id},
        ).fetchall()

        carbon = conn.execute(
            text(
                """SELECT class_code, class_name, area_ha, coef_t_per_ha, carbon_t, co2eq_t
                   FROM fcd_carbon_stats
                   WHERE run_id = :r AND aoi_id = :a
                   ORDER BY class_code"""
            ),
            {"r": run_id, "a": aoi_id},
        ).fetchall()

        run = conn.execute(
            text(
                "SELECT year, start_date, end_date, algorithm_version "
                "FROM fcd_runs WHERE id = :r"
            ),
            {"r": run_id},
        ).fetchone()

    return {
        "aoi":     dict(aoi._mapping)    if aoi    else {},
        "run":     dict(run._mapping)    if run    else {},
        "classes": [dict(r._mapping) for r in classes],
        "carbon":  [dict(r._mapping) for r in carbon],
    }


def _build_excel(data: dict, path: Path) -> None:
    wb = openpyxl.Workbook()
    aoi = data["aoi"]
    run = data["run"]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1a5c2a")

    # ── Sheet 1: Summary ──────────────────────────────────────────
    ws = wb.active
    ws.title = "FCD Summary"
    ws.append(["FCD Tamil Nadu Platform — District Report"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["AOI Name",          aoi.get("name", "")])
    ws.append(["Unit Type",         aoi.get("unit_type", "")])
    ws.append(["Year",              run.get("year", "")])
    ws.append(["Date Range",        f"{run.get('start_date','')} \u2192 {run.get('end_date','')}"])
    ws.append(["Algorithm Version", run.get("algorithm_version", "")])
    ws.append([])

    headers = ["Class Code", "Class Name", "Area (ha)", "% Area"]
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in data["classes"]:
        ws.append(
            [row["class_code"], row["class_name"],
             float(row["area_ha"]), float(row["percent_area"] or 0)]
        )
        ws.cell(ws.max_row, 2).fill = PatternFill(
            "solid", fgColor=CLASS_COLORS.get(row["class_code"], "cccccc")
        )

    for col, width in zip(["A", "B", "C", "D"], [12, 20, 16, 10]):
        ws.column_dimensions[col].width = width

    # ── Sheet 2: Carbon ───────────────────────────────────────────
    wc = wb.create_sheet("Carbon Estimates")
    wc.append(["Class Code", "Class Name", "Area (ha)", "Coef (t/ha)", "Carbon (t)", "CO\u2082-eq (t)"])
    for cell in wc[1]:
        cell.font = header_font
        cell.fill = header_fill
    for row in data["carbon"]:
        wc.append(
            [row["class_code"], row["class_name"],
             float(row["area_ha"]), float(row["coef_t_per_ha"]),
             float(row["carbon_t"]), float(row["co2eq_t"])]
        )

    # ── Sheet 3: Bar chart ────────────────────────────────────────
    chart_ws = wb.create_sheet("Chart")
    chart_ws.append(["Class", "Area (ha)"])
    for row in data["classes"]:
        chart_ws.append([row["class_name"], float(row["area_ha"])])

    chart = BarChart()
    chart.title    = f"FCD Class Distribution \u2014 {aoi.get('name','')} {run.get('year','')}"
    chart.y_axis.title = "Area (ha)"
    chart.x_axis.title = "FCD Class"
    data_ref = Reference(chart_ws, min_col=2, min_row=1, max_row=chart_ws.max_row)
    cats_ref = Reference(chart_ws, min_col=1, min_row=2, max_row=chart_ws.max_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.shape = 4
    chart_ws.add_chart(chart, "D2")

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _build_pdf(data: dict, path: Path) -> None:
    aoi = data["aoi"]
    run = data["run"]

    rows_classes = "".join(
        f"""<tr>
          <td style="background:{_hex(c['class_code'])};color:{'#fff' if c['class_code'] in (1,2) else '#000'};
              padding:6px 10px;border-radius:3px">{c['class_name']}</td>
          <td>{float(c['area_ha']):,.2f}</td>
          <td>{float(c['percent_area'] or 0):.2f}%</td>
        </tr>"""
        for c in data["classes"]
    )

    rows_carbon = "".join(
        f"<tr><td>{c['class_name']}</td><td>{float(c['area_ha']):,.2f}</td>"
        f"<td>{float(c['coef_t_per_ha'])}</td>"
        f"<td>{float(c['carbon_t']):,.2f}</td><td>{float(c['co2eq_t']):,.2f}</td></tr>"
        for c in data["carbon"]
        if float(c["carbon_t"]) > 0
    )

    total_carbon = sum(float(c["carbon_t"]) for c in data["carbon"])
    total_co2    = sum(float(c["co2eq_t"])  for c in data["carbon"])

    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
  @page {{ size: A4; margin: 2cm; }}
  body {{ font-family: Arial, sans-serif; font-size: 11pt; color: #222; }}
  h1   {{ color: #1a5c2a; font-size: 18pt; margin-bottom: 4px; }}
  h2   {{ color: #1a5c2a; font-size: 13pt; border-bottom: 2px solid #1a5c2a; padding-bottom: 4px; }}
  .meta {{ background: #f4f9f4; border-left: 4px solid #1a5c2a; padding: 10px 16px; margin: 12px 0; }}
  table {{ border-collapse: collapse; width: 100%; margin-top: 8px; }}
  th    {{ background: #1a5c2a; color: white; padding: 7px 10px; text-align: left; font-size: 10pt; }}
  td    {{ padding: 6px 10px; border-bottom: 1px solid #ddd; font-size: 10pt; }}
  tr:nth-child(even) td {{ background: #f9f9f9; }}
  .footer {{ margin-top: 30px; font-size: 9pt; color: #888; text-align: center; }}
</style>
</head><body>
<h1>\U0001f33f Forest Canopy Density Report</h1>
<div class="meta">
  <strong>Area:</strong> {aoi.get('name','N/A')} &nbsp;|&nbsp;
  <strong>Type:</strong> {aoi.get('unit_type','').replace('_',' ').title()} &nbsp;|&nbsp;
  <strong>Year:</strong> {run.get('year','N/A')} &nbsp;|&nbsp;
  <strong>Date Range:</strong> {run.get('start_date','')} \u2192 {run.get('end_date','')}
</div>

<h2>FCD Class Distribution</h2>
<table>
  <tr><th>Class</th><th>Area (ha)</th><th>Coverage</th></tr>
  {rows_classes}
</table>

<h2>Carbon Sequestration Estimates</h2>
<table>
  <tr><th>Class</th><th>Area (ha)</th><th>Coef (t/ha)</th><th>Carbon (t)</th><th>CO\u2082-eq (t)</th></tr>
  {rows_carbon}
  <tr style="font-weight:bold;background:#e8f5e9">
    <td colspan="3">Total</td>
    <td>{total_carbon:,.2f}</td>
    <td>{total_co2:,.2f}</td>
  </tr>
</table>

<div class="footer">
  Generated by FCD Tamil Nadu Platform &middot; Sentinel-2 SR &middot;
  Algorithm v{run.get('algorithm_version','1.0.0')}
  &nbsp;&middot;&nbsp; &copy; Athithiyan M R, Auroville Consulting
</div>
</body></html>"""

    path.parent.mkdir(parents=True, exist_ok=True)
    HTML(string=html).write_pdf(str(path))


@shared_task(bind=True, name="tasks.generate_report", max_retries=3, default_retry_delay=30)
def generate_report(self, job_id: str, run_id: str, aoi_id: str, report_type: str):
    with engine.connect() as conn:
        conn.execute(
            text("UPDATE report_jobs SET status='running' WHERE id=:id"),
            {"id": job_id},
        )
        conn.commit()

    try:
        data      = _fetch_stats(run_id, aoi_id)
        aoi_slug  = data["aoi"].get("name", aoi_id).replace(" ", "_")
        year      = data["run"].get("year", "unknown")
        base_name = f"{report_type}_{aoi_slug}_{year}"

        xlsx_path = REPORT_DIR / f"{base_name}.xlsx"
        pdf_path  = REPORT_DIR / f"{base_name}.pdf"

        _build_excel(data, xlsx_path)
        _build_pdf(data,   pdf_path)

        xlsx_url = f"{BASE_URL}/static/reports/{xlsx_path.name}"
        pdf_url  = f"{BASE_URL}/static/reports/{pdf_path.name}"

        with engine.connect() as conn:
            conn.execute(
                text(
                    """UPDATE report_jobs
                       SET status='completed', pdf_url=:pdf, xlsx_url=:xlsx,
                           completed_at=now()
                       WHERE id=:id"""
                ),
                {"pdf": pdf_url, "xlsx": xlsx_url, "id": job_id},
            )
            conn.commit()

    except Exception as exc:
        with engine.connect() as conn:
            conn.execute(
                text("UPDATE report_jobs SET status='failed' WHERE id=:id"),
                {"id": job_id},
            )
            conn.commit()
        raise self.retry(exc=exc)
